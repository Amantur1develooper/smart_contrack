import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

# Создание данных
data = {
    'Дни недели': ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'],
    'Доход (сом)': [3245.20, 4572.50, 6251.66, 2125.20, 3896.60, 5420.30, 6050.60],
    'Расход (сом)': [3628.50, 5320.50, 5292.10, 3824.30, 3020.10, 4262.10, 4369.50]
}

df = pd.DataFrame(data)

# Вычисления
df['Финансовый результат (сом)'] = df['Доход (сом)'] - df['Расход (сом)']
average_income = df['Доход (сом)'].mean()
average_expense = df['Расход (сом)'].mean()
max_income = df['Доход (сом)'].max()
min_expense = df['Расход (сом)'].min()
total_financial_result = df['Финансовый результат (сом)'].sum()

# Добавление итогов в таблицу
summary_data = pd.DataFrame({
    'Дни недели': ['Среднее значение', 'Максимальный доход', 'Минимальный расход', 'Общий финансовый результат за неделю'],
    'Доход (сом)': [average_income, max_income, '', ''],
    'Расход (сом)': [average_expense, '', min_expense, ''],
    'Финансовый результат (сом)': ['', '', '', total_financial_result]
})

# Создание финальной таблицы
final_df = pd.concat([df, summary_data], ignore_index=True)

# Создание книги и листов Excel
wb = Workbook()
ws1 = wb.active
ws1.title = "Расчеты"

# Запись данных в первый лист
for r in dataframe_to_rows(final_df, index=False, header=True):
    ws1.append(r)

# Копирование данных на второй лист
ws2 = wb.create_sheet(title="Отбор данных")
for r in dataframe_to_rows(final_df, index=False, header=True):
    ws2.append(r)

# Построение диаграммы
chart = LineChart()
data = Reference(ws1, min_col=4, min_row=1, max_col=4, max_row=8)
categories = Reference(ws1, min_col=1, min_row=2, max_row=8)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Изменение финансовых результатов"
chart.x_axis.title = "Дни недели"
chart.y_axis.title = "Финансовый результат (сом)"

ws3 = wb.create_sheet(title="Диаграмма")
ws3.add_chart(chart, "A1")

# Сохранение в файл
wb.save("Финансовый расчет.xlsx")
